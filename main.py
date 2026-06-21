import os
import sys
from datetime import datetime
from collections import defaultdict
from colorama import init, Fore, Style, Back
from tabulate import tabulate

from config import (
    SAMPLES_DIR, RULE_CATEGORIES, DEAL_STATUS,
    LOW_SCORE_THRESHOLD, BASE_DIR, EXPORTS_DIR
)
from importer import RecordingImporter
from batch_processor import BatchProcessor
from analyzer import ResultAnalyzer, RecordManager, ExcelExporter, BatchComparator, WeeklyAnalyzer
from qa_engine import QARuleEngine

init(autoreset=True)


class QACLI:
    def __init__(self):
        self.importer = RecordingImporter()
        self.processor = BatchProcessor()
        self.analyzer = ResultAnalyzer()
        self.record_manager = RecordManager()
        self.exporter = ExcelExporter()
        self.comparator = BatchComparator(self.analyzer)
        self.weekly_analyzer = WeeklyAnalyzer(self.record_manager, self.analyzer)

        self.current_dir = SAMPLES_DIR
        self.transcript_dir = None
        self.use_simulator_fallback = True
        self.current_category = "light"
        self.current_deal_filter = "all"
        self.current_batch = None
        self.current_batch_name = ""

    def clear_screen(self):
        os.system("cls" if os.name == "nt" else "clear")

    def print_header(self):
        print(f"{Fore.CYAN}{Style.BRIGHT}")
        print("=" * 60)
        print("      医美集团录音批量质检工具 v1.0")
        print("=" * 60)
        print(Style.RESET_ALL)

    def print_menu(self):
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 主菜单 】{Style.RESET_ALL}")
        print(f"\n{Fore.GREEN}[1] 导入录音{Style.RESET_ALL}    指定文件夹，解析门店和咨询师")
        print(f"{Fore.GREEN}[2] 规则选择{Style.RESET_ALL}    轻医美/手术类/皮肤类，成交筛选")
        print(f"{Fore.GREEN}[3] 开始质检{Style.RESET_ALL}    批处理，实时显示进度")
        print(f"{Fore.GREEN}[4] 异常清单{Style.RESET_ALL}    低分录音、问题明细（分页浏览）")
        print(f"{Fore.GREEN}[5] 门店汇总{Style.RESET_ALL}    各门店得分、Top10问题")
        print(f"{Fore.GREEN}[6] 结果导出{Style.RESET_ALL}    导出表格给主管复核")
        print(f"{Fore.GREEN}[7] 历史记录{Style.RESET_ALL}    查看历史批次，对比分析")
        print(f"{Fore.GREEN}[8] 周会模式{Style.RESET_ALL}    按周汇总趋势，导出周会包")
        print(f"{Fore.GREEN}[9] 咨询师下钻{Style.RESET_ALL}  选咨询师看走势和复听样本")
        print(f"{Fore.RED}[0] 退出{Style.RESET_ALL}")

    def print_status_bar(self):
        cat_name = RULE_CATEGORIES.get(self.current_category, self.current_category)
        deal_name = DEAL_STATUS.get(self.current_deal_filter, self.current_deal_filter)

        file_count = len(self.importer.list_recording_files(self.current_dir))

        print(f"\n{Fore.CYAN}{'-' * 60}")
        print(f"  当前文件夹: {self.current_dir}")
        if self.transcript_dir:
            print(f"  转写文件夹: {Fore.GREEN}{self.transcript_dir}{Style.RESET_ALL}")
        else:
            print(f"  转写文件夹: {Fore.LIGHTBLACK_EX}未设置（使用同名txt或模拟）{Style.RESET_ALL}")
        print(f"  录音文件数: {file_count} 个")
        print(f"  质检规则: {cat_name} | 成交筛选: {deal_name}")
        if self.current_batch:
            print(f"  当前批次: {self.current_batch_name} ({self.current_batch.batch_id})")
            print(f"  批处理状态: 已完成 | 通过率: {self.current_batch.passed_count}/{self.current_batch.total_count}")
        else:
            print(f"  当前批次: 无")
        print(f"{'-' * 60}{Style.RESET_ALL}")

    def get_input(self, prompt: str, default: str = "") -> str:
        if default:
            prompt = f"{prompt} [{default}]: "
        else:
            prompt = f"{prompt}: "

        try:
            value = input(f"{Fore.WHITE}{prompt}{Style.RESET_ALL}").strip()
            return value if value else default
        except (EOFError, KeyboardInterrupt):
            return default

    def pause(self):
        input(f"\n{Fore.YELLOW}按回车键继续...{Style.RESET_ALL}")

    # ===== 操作1: 导入录音 =====
    def action_import(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 导入录音 】{Style.RESET_ALL}")
        print(f"  请指定录音文件夹路径")
        print(f"  文件名格式: 门店_咨询师_日期_时间_成交状态.wav")
        print(f"  例: 朝阳店_张美丽_20240115_143000_dealt.wav\n")

        default_dir = self.current_dir or SAMPLES_DIR
        directory = self.get_input("请输入录音文件夹路径", default_dir)

        if not os.path.exists(directory):
            print(f"\n{Fore.RED}错误: 文件夹不存在 - {directory}{Style.RESET_ALL}")
            self.pause()
            return

        files = self.importer.list_recording_files(directory)
        if not files:
            print(f"\n{Fore.RED}该文件夹下没有找到 .wav 录音文件{Style.RESET_ALL}")
            self.pause()
            return

        self.current_dir = directory

        print(f"\n{Fore.GREEN}✓ 找到 {len(files)} 个录音文件{Style.RESET_ALL}")

        transcript_dir_default = self.transcript_dir if self.transcript_dir else "留空表示不设置"
        transcript_input = self.get_input("\n请输入转写稿文件夹路径（留空则使用同名txt或模拟）", transcript_dir_default)
        if transcript_input and transcript_input != "留空表示不设置":
            if os.path.exists(transcript_input):
                self.transcript_dir = transcript_input
                self.importer.set_transcript_dir(transcript_input)
                print(f"{Fore.GREEN}✓ 已设置转写文件夹{Style.RESET_ALL}")
            else:
                print(f"{Fore.YELLOW}⚠ 转写文件夹不存在，继续使用默认设置{Style.RESET_ALL}")
                self.transcript_dir = None
                self.importer.set_transcript_dir(None)
        else:
            self.transcript_dir = None
            self.importer.set_transcript_dir(None)

        fallback_choice = self.get_input("无转写稿时是否使用模拟数据？(y/n)", "y" if self.use_simulator_fallback else "n")
        self.use_simulator_fallback = fallback_choice.lower() == "y"
        self.importer.set_simulator_fallback(self.use_simulator_fallback)

        self.importer.import_from_directory(self.current_dir, self.current_category, self.current_deal_filter)
        stats = self.importer.get_transcript_stats()

        print(f"\n{Style.BRIGHT}转写稿统计:{Style.RESET_ALL}")
        print(f"  真实转写: {Fore.GREEN}{stats['real']} 个{Style.RESET_ALL}")
        if stats['simulated'] > 0:
            print(f"  模拟转写: {Fore.YELLOW}{stats['simulated']} 个{Style.RESET_ALL}")
        if stats['failed'] > 0:
            print(f"  导入失败: {Fore.RED}{stats['failed']} 个{Style.RESET_ALL}")

        missing_info = self.importer.check_missing_transcripts(self.current_dir)
        missing = missing_info["missing_txt"]
        orphan = missing_info["orphan_txt"]

        print(f"\n{Style.BRIGHT}转写匹配检查:{Style.RESET_ALL}")
        print(f"  录音文件: {missing_info['total_wav']} 个")
        print(f"  转写文件: {missing_info['total_txt']} 个")
        print(f"  成功匹配: {Fore.GREEN}{missing_info['matched_count']} 个{Style.RESET_ALL}")
        print(f"  缺少转写: {Fore.YELLOW}{len(missing)} 个录音{Style.RESET_ALL}")
        if orphan:
            print(f"  孤立转写: {Fore.LIGHTBLACK_EX}{len(orphan)} 个txt无对应录音{Style.RESET_ALL}")

        if missing:
            print(f"\n{Style.BRIGHT}📋 缺少转写的录音（前10个）:{Style.RESET_ALL}")
            for item in missing[:10]:
                print(f"  {Fore.YELLOW}• {item['filename']}{Style.RESET_ALL}")
                print(f"    {Fore.LIGHTBLACK_EX}{item['full_path']}{Style.RESET_ALL}")
            if len(missing) > 10:
                print(f"  ... 还有 {len(missing) - 10} 个")

        if orphan:
            print(f"\n{Style.BRIGHT}📋 未匹配的转写稿（前5个）:{Style.RESET_ALL}")
            for item in orphan[:5]:
                print(f"  {Fore.LIGHTBLACK_EX}• {item['filename']}{Style.RESET_ALL}")

        if missing or orphan:
            export_missing = self.get_input("\n导出转写缺失清单给数据同事？(y/n)", "n")
            if export_missing.lower() == "y":
                self._export_missing_list(missing_info)

        sample_files = files[:8]
        table_data = []
        for fname in sample_files:
            parsed = self.importer.parser.parse(fname)
            if parsed:
                deal = "已成交" if parsed["deal"] == "dealt" else "未成交"
                txt_name = os.path.splitext(fname)[0] + ".txt"
                has_txt = os.path.exists(os.path.join(self.current_dir, txt_name))
                if self.transcript_dir:
                    has_txt = has_txt or os.path.exists(os.path.join(self.transcript_dir, txt_name))
                source = "真实转写" if has_txt else ("模拟转写" if self.use_simulator_fallback else "无转写")
                source_color = Fore.GREEN if has_txt else (Fore.YELLOW if self.use_simulator_fallback else Fore.RED)
                table_data.append([
                    parsed["store"],
                    parsed["consultant"],
                    parsed["date"],
                    deal,
                    f"{source_color}{source}{Style.RESET_ALL}"
                ])
            else:
                table_data.append(["(格式不符)", fname, "", "", ""])

        print(f"\n{Style.BRIGHT}文件预览（前8个）:{Style.RESET_ALL}\n")
        print(tabulate(table_data, headers=["门店", "咨询师", "日期", "成交状态", "转写来源"], tablefmt="simple"))

        if len(files) > 8:
            print(f"\n  ... 还有 {len(files) - 8} 个文件")

        self.pause()

    def _export_missing_list(self, missing_info: dict):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "总览"
        ws1["A1"] = "转写稿缺失清单"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1.merge_cells("A1:D1")

        rows = [
            ["录音文件总数", missing_info["total_wav"]],
            ["转写文件总数", missing_info["total_txt"]],
            ["成功匹配数", missing_info["matched_count"]],
            ["缺少转写的录音数", len(missing_info["missing_txt"])],
            ["无对应录音的转写数", len(missing_info["orphan_txt"])],
        ]
        for i, (key, val) in enumerate(rows, 3):
            ws1.cell(row=i, column=1, value=key)
            ws1.cell(row=i, column=2, value=val)

        ws2 = wb.create_sheet("缺失明细")
        headers = ["序号", "文件名", "完整路径", "类型"]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h).font = Font(bold=True)

        row_idx = 2
        for item in missing_info["missing_txt"]:
            ws2.cell(row=row_idx, column=1, value=row_idx - 1)
            ws2.cell(row=row_idx, column=2, value=item["filename"])
            ws2.cell(row=row_idx, column=3, value=item["full_path"])
            ws2.cell(row=row_idx, column=4, value="缺txt")
            row_idx += 1

        for item in missing_info["orphan_txt"]:
            ws2.cell(row=row_idx, column=1, value=row_idx - 1)
            ws2.cell(row=row_idx, column=2, value=item["filename"])
            ws2.cell(row=row_idx, column=3, value=item["full_path"])
            cell = ws2.cell(row=row_idx, column=4, value="多出的txt")
            cell.font = Font(color="FF0000")
            row_idx += 1

        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 60
        ws2.column_dimensions["D"].width = 12

        os.makedirs(EXPORTS_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"转写缺失清单_{timestamp}.xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)
        wb.save(filepath)

        print(f"\n{Fore.GREEN}✓ 已导出缺失清单{Style.RESET_ALL}")
        print(f"  {Fore.CYAN}{filepath}{Style.RESET_ALL}")
        print(f"  缺txt录音: {len(missing_info['missing_txt'])} 个 | 多出txt: {len(missing_info['orphan_txt'])} 个")
        print(f"  请将此文件发给数据同事补充转写")

    def _export_review_task_list(self, store: str, consultant: str, low_recordings: list, batches: list):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        os.makedirs(EXPORTS_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"复听任务清单_{store}_{consultant}_{timestamp}.xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        ws1 = wb.active
        ws1.title = "按问题类型分组"
        ws1.merge_cells("A1:H1")
        ws1["A1"] = f"复听任务清单 - {store} - {consultant}"
        ws1["A1"].font = Font(bold=True, size=14, color="1F4E78")

        headers = ["分组", "#", "分数", "批次", "问题类型", "文件名", "完整路径", "已复核"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        by_issue = defaultdict(list)
        for lr in low_recordings:
            for issue in lr["issues"]:
                by_issue[issue].append(lr)

        row = 4
        idx = 1
        for issue, records in sorted(by_issue.items(), key=lambda x: len(x[1]), reverse=True):
            sorted_records = sorted(records, key=lambda x: x["score"])
            for i, lr in enumerate(sorted_records):
                ws1.cell(row=row, column=1, value=issue if i == 0 else "")
                ws1.cell(row=row, column=2, value=idx)
                score_cell = ws1.cell(row=row, column=3, value=lr["score"])
                score_cell.font = Font(color="C00000", bold=True)
                ws1.cell(row=row, column=4, value=lr["batch_name"])
                ws1.cell(row=row, column=5, value="; ".join(lr["issues"]))
                ws1.cell(row=row, column=6, value=lr["file_name"])
                ws1.cell(row=row, column=7, value=lr["file_path"])
                ws1.cell(row=row, column=8, value="")
                row += 1
                idx += 1

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 6
        ws1.column_dimensions["C"].width = 8
        ws1.column_dimensions["D"].width = 18
        ws1.column_dimensions["E"].width = 30
        ws1.column_dimensions["F"].width = 40
        ws1.column_dimensions["G"].width = 60
        ws1.column_dimensions["H"].width = 10

        ws2 = wb.create_sheet("按批次分组")
        ws2.merge_cells("A1:H1")
        ws2["A1"] = f"复听任务清单 - 按批次分组 - {store} - {consultant}"
        ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")

        headers2 = ["分组", "#", "分数", "问题类型", "文件名", "完整路径", "已复核"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 4
        idx = 1
        by_batch = defaultdict(list)
        for lr in low_recordings:
            by_batch[lr["batch_name"]].append(lr)
        for batch in batches:
            if batch.batch_name in by_batch:
                records = sorted(by_batch[batch.batch_name], key=lambda x: x["score"])
                for i, lr in enumerate(records):
                    ws2.cell(row=row, column=1, value=batch.batch_name if i == 0 else "")
                    ws2.cell(row=row, column=2, value=idx)
                    score_cell = ws2.cell(row=row, column=3, value=lr["score"])
                    score_cell.font = Font(color="C00000", bold=True)
                    ws2.cell(row=row, column=4, value="; ".join(lr["issues"]))
                    ws2.cell(row=row, column=5, value=lr["file_name"])
                    ws2.cell(row=row, column=6, value=lr["file_path"])
                    ws2.cell(row=row, column=7, value="")
                    row += 1
                    idx += 1

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 6
        ws2.column_dimensions["C"].width = 8
        ws2.column_dimensions["D"].width = 30
        ws2.column_dimensions["E"].width = 40
        ws2.column_dimensions["F"].width = 60
        ws2.column_dimensions["G"].width = 10

        ws3 = wb.create_sheet("按分数排序")
        ws3.merge_cells("A1:G1")
        ws3["A1"] = f"复听任务清单 - 按分数排序 - {store} - {consultant}"
        ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")

        headers3 = ["#", "分数", "批次", "问题类型", "文件名", "完整路径", "已复核"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 4
        sorted_by_score = sorted(low_recordings, key=lambda x: x["score"])
        for i, lr in enumerate(sorted_by_score, 1):
            ws3.cell(row=row, column=1, value=i)
            score_cell = ws3.cell(row=row, column=2, value=lr["score"])
            score_cell.font = Font(color="C00000", bold=True)
            ws3.cell(row=row, column=3, value=lr["batch_name"])
            ws3.cell(row=row, column=4, value="; ".join(lr["issues"]))
            ws3.cell(row=row, column=5, value=lr["file_name"])
            ws3.cell(row=row, column=6, value=lr["file_path"])
            ws3.cell(row=row, column=7, value="")
            row += 1

        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 8
        ws3.column_dimensions["C"].width = 18
        ws3.column_dimensions["D"].width = 30
        ws3.column_dimensions["E"].width = 40
        ws3.column_dimensions["F"].width = 60
        ws3.column_dimensions["G"].width = 10

        wb.save(filepath)
        return filepath

    # ===== 操作2: 规则选择 =====
    def action_rules(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 规则选择 】{Style.RESET_ALL}\n")

        print(f"{Style.BRIGHT}请选择质检规则类型:{Style.RESET_ALL}")
        categories = list(RULE_CATEGORIES.items())
        for i, (key, name) in enumerate(categories, 1):
            marker = " ◄" if key == self.current_category else ""
            print(f"  [{i}] {name}{marker}")

        cat_choice = self.get_input("\n请选择规则编号", "")
        if cat_choice.isdigit() and 1 <= int(cat_choice) <= len(categories):
            self.current_category = categories[int(cat_choice) - 1][0]

        print(f"\n{Style.BRIGHT}请选择成交筛选:{Style.RESET_ALL}")
        deals = list(DEAL_STATUS.items())
        for i, (key, name) in enumerate(deals, 1):
            marker = " ◄" if key == self.current_deal_filter else ""
            print(f"  [{i}] {name}{marker}")

        deal_choice = self.get_input("\n请选择筛选编号", "")
        if deal_choice.isdigit() and 1 <= int(deal_choice) <= len(deals):
            self.current_deal_filter = deals[int(deal_choice) - 1][0]

        engine = QARuleEngine(category=self.current_category)
        ban_words = engine.get_ban_words()

        print(f"\n{Style.BRIGHT}当前规则配置:{Style.RESET_ALL}")
        print(f"  规则类型: {RULE_CATEGORIES[self.current_category]}")
        print(f"  成交筛选: {DEAL_STATUS[self.current_deal_filter]}")
        print(f"  禁用词库（{len(ban_words)}个）: {', '.join(ban_words[:5])}...")
        print(f"  打断阈值: {__import__('config').INTERRUPTION_THRESHOLD} 次")
        print(f"  低分阈值: {LOW_SCORE_THRESHOLD} 分")

        self.pause()

    # ===== 操作3: 开始质检 =====
    def action_batch(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 开始质检 】{Style.RESET_ALL}\n")

        if not self.current_dir or not os.path.exists(self.current_dir):
            print(f"{Fore.RED}请先指定录音文件夹（选择 导入录音）{Style.RESET_ALL}")
            self.pause()
            return

        default_name = f"周检_{datetime.now().strftime('%Y%m%d')}"
        self.current_batch_name = self.get_input("请输入批次名称", default_name)

        cat_name = RULE_CATEGORIES.get(self.current_category, self.current_category)
        deal_name = DEAL_STATUS.get(self.current_deal_filter, self.current_deal_filter)

        print(f"\n  质检规则: {cat_name}")
        print(f"  成交筛选: {deal_name}")
        print(f"  源文件夹: {self.current_dir}")

        confirm = self.get_input("\n确认开始质检？(y/n)", "y")
        if confirm.lower() != "y":
            print("已取消")
            self.pause()
            return

        print(f"\n{Fore.CYAN}正在加载录音文件...{Style.RESET_ALL}")

        try:
            self.processor.importer.set_transcript_dir(self.transcript_dir)
            self.processor.importer.set_simulator_fallback(self.use_simulator_fallback)

            batch = self.processor.process_directory(
                name=self.current_batch_name,
                directory=self.current_dir,
                category=self.current_category,
                deal_filter=self.current_deal_filter
            )
            self.current_batch = batch

            if batch.total_count == 0:
                print(f"\n{Fore.RED}没有找到符合条件的录音文件{Style.RESET_ALL}")
                self.pause()
                return

            self.record_manager.save_batch(batch)
            transcript_stats = self.processor.importer.get_transcript_stats()

            print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ 质检完成！{Style.RESET_ALL}")
            print(f"\n  批次ID: {batch.batch_id}")
            print(f"  录音总数: {batch.total_count}")
            print(f"  通过数: {batch.passed_count}  ({batch.passed_count / batch.total_count * 100:.1f}%)")
            print(f"  异常数: {batch.failed_count}  ({batch.failed_count / batch.total_count * 100:.1f}%)")
            print(f"  平均得分: {sum(r.total_score for r in batch.results) / batch.total_count:.1f}")
            print(f"\n  转写来源:")
            print(f"    真实转写: {Fore.GREEN}{transcript_stats['real']} 个{Style.RESET_ALL}")
            if transcript_stats['simulated'] > 0:
                print(f"    模拟转写: {Fore.YELLOW}{transcript_stats['simulated']} 个{Style.RESET_ALL}")
            if transcript_stats['failed'] > 0:
                print(f"    导入失败: {Fore.RED}{transcript_stats['failed']} 个{Style.RESET_ALL}")

            duration = (batch.end_time - batch.start_time).total_seconds() if batch.end_time else 0
            print(f"  耗时: {duration:.1f} 秒")

        except Exception as e:
            print(f"\n{Fore.RED}质检过程出错: {e}{Style.RESET_ALL}")

        self.pause()

    # ===== 操作4: 异常清单 =====
    def action_issues(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 异常清单 】{Style.RESET_ALL}\n")

        if not self.current_batch:
            print(f"{Fore.RED}请先执行批处理（选择 开始质检）{Style.RESET_ALL}")
            self.pause()
            return

        low_scores = self.analyzer.get_low_score_list(self.current_batch)

        print(f"{Style.BRIGHT}低分录音（<{LOW_SCORE_THRESHOLD}分）共 {len(low_scores)} 个{Style.RESET_ALL}")
        print(f"{Fore.LIGHTBLACK_EX}提示: 输入序号查看详情，n=下一页，p=上一页，q=返回{Style.RESET_ALL}\n")

        if not low_scores:
            print(f"{Fore.GREEN}太棒了！本次质检没有低分录音{Style.RESET_ALL}")
        else:
            PAGE_SIZE = 10
            total_pages = (len(low_scores) + PAGE_SIZE - 1) // PAGE_SIZE
            current_page = 0

            while True:
                self.clear_screen()
                self.print_header()
                print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 异常清单 - 低分录音 】{Style.RESET_ALL}")
                print(f"  共 {len(low_scores)} 条，第 {current_page + 1}/{total_pages} 页\n")

                start = current_page * PAGE_SIZE
                end = min(start + PAGE_SIZE, len(low_scores))
                page_items = low_scores[start:end]

                for i, result in enumerate(page_items, start + 1):
                    rec = result.recording
                    score_color = Fore.RED if result.total_score < 40 else Fore.YELLOW
                    deal = "✓已成交" if rec.is_dealt else "○未成交"
                    source = "真实转写" if getattr(rec, "transcript_source", "simulated") == "real" else "模拟转写"
                    source_color = Fore.GREEN if getattr(rec, "transcript_source", "simulated") == "real" else Fore.LIGHTBLACK_EX

                    print(f"  [{i:2d}] {score_color}{result.total_score:3d}分{Style.RESET_ALL}  "
                          f"{rec.store} - {rec.consultant}  {deal}  {source_color}{source}{Style.RESET_ALL}")
                    print(f"       {Fore.CYAN}{rec.file_path}{Style.RESET_ALL}")
                    if result.issues:
                        issues_str = "; ".join(result.issues[:2])
                        if len(result.issues) > 2:
                            issues_str += f" ...(+{len(result.issues) - 2})"
                        print(f"       {Fore.LIGHTBLACK_EX}问题: {issues_str}{Style.RESET_ALL}")
                    print()

                print(f"{Fore.LIGHTBLACK_EX}操作: 序号=查看详情 | n=下一页 | p=上一页 | q=返回{Style.RESET_ALL}")
                choice = self.get_input("请输入操作", "q")

                if choice.lower() == 'q':
                    break
                elif choice.lower() == 'n':
                    if current_page < total_pages - 1:
                        current_page += 1
                elif choice.lower() == 'p':
                    if current_page > 0:
                        current_page -= 1
                elif choice.isdigit():
                    idx = int(choice)
                    if 1 <= idx <= len(low_scores):
                        self._show_detail(low_scores[idx - 1])
                        input(f"\n{Fore.YELLOW}按回车键继续浏览...{Style.RESET_ALL}")

        print(f"\n{Style.BRIGHT}问题类型统计 Top10:{Style.RESET_ALL}\n")
        top_issues = self.analyzer.get_top_issues(self.current_batch, top_n=10)
        table_data = []
        for i, (issue, count) in enumerate(top_issues, 1):
            pct = count / self.current_batch.total_count * 100
            table_data.append([i, issue, count, f"{pct:.1f}%"])

        print(tabulate(table_data, headers=["排名", "问题类型", "出现次数", "占比"], tablefmt="simple"))

        self.pause()

    def _show_detail(self, result):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 质检详情 】{Style.RESET_ALL}\n")

        rec = result.recording
        print(f"  文件: {Fore.CYAN}{rec.file_name}{Style.RESET_ALL}")
        print(f"  路径: {rec.file_path}")
        print(f"  门店: {rec.store} | 咨询师: {rec.consultant}")
        print(f"  日期: {rec.record_date} {rec.record_time}")
        print(f"  状态: {'已成交' if rec.is_dealt else '未成交'} | 时长: {rec.duration_seconds}秒")
        print(f"\n  {Style.BRIGHT}总得分: {Fore.RED if result.is_low_score else Fore.GREEN}{result.total_score}分{Style.RESET_ALL}\n")

        print(f"{Style.BRIGHT}各分项得分:{Style.RESET_ALL}")
        for check in result.check_results:
            status = f"{Fore.GREEN}✓{Style.RESET_ALL}" if check.passed else f"{Fore.RED}✗{Style.RESET_ALL}"
            print(f"  {status} {check.rule_name}: {check.score}分 - {check.detail}")
            if check.evidence:
                for ev in check.evidence[:2]:
                    print(f"      证据: {Fore.LIGHTBLACK_EX}{ev}{Style.RESET_ALL}")

        print(f"\n{Style.BRIGHT}问题列表:{Style.RESET_ALL}")
        for i, issue in enumerate(result.issues, 1):
            print(f"  {i}. {Fore.RED}{issue}{Style.RESET_ALL}")

        print(f"\n{Style.BRIGHT}转写文本预览:{Style.RESET_ALL}")
        transcript = rec.transcript
        if len(transcript) > 300:
            transcript = transcript[:300] + "..."
        print(f"  {Fore.LIGHTBLACK_EX}{transcript}{Style.RESET_ALL}")

    # ===== 操作5: 门店汇总 =====
    def action_stores(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 门店汇总 】{Style.RESET_ALL}\n")

        if not self.current_batch:
            print(f"{Fore.RED}请先执行批处理（选择 开始质检）{Style.RESET_ALL}")
            self.pause()
            return

        summaries = self.analyzer.get_store_summaries(self.current_batch)

        print(f"{Style.BRIGHT}各门店质检得分排名（由低到高）:{Style.RESET_ALL}\n")

        table_data = []
        for i, summary in enumerate(summaries, 1):
            color = Fore.RED if summary.avg_score < LOW_SCORE_THRESHOLD else Fore.GREEN
            top_issues_str = "; ".join([f"{issue}({count})" for issue, count in summary.top_issues[:2]])
            table_data.append([
                i,
                summary.store_name,
                summary.total_count,
                f"{color}{summary.avg_score}{Style.RESET_ALL}",
                f"{summary.pass_rate}%",
                summary.issue_count,
                top_issues_str
            ])

        print(tabulate(
            table_data,
            headers=["排名", "门店", "录音数", "平均分", "通过率", "问题数", "主要问题"],
            tablefmt="simple"
        ))

        print(f"\n{Style.BRIGHT}全店问题 Top 10:{Style.RESET_ALL}\n")
        top_issues = self.analyzer.get_top_issues(self.current_batch, top_n=10)
        table_data = []
        for i, (issue, count) in enumerate(top_issues, 1):
            pct = count / self.current_batch.total_count * 100
            bar_len = int(pct / 2)
            bar = "█" * bar_len
            table_data.append([i, issue, count, f"{pct:.1f}%", bar])

        print(tabulate(table_data, headers=["排名", "问题类型", "次数", "占比", "分布"], tablefmt="simple"))

        self.pause()

    # ===== 操作6: 结果导出 =====
    def action_export(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 结果导出 】{Style.RESET_ALL}\n")

        if not self.current_batch:
            print(f"{Fore.RED}请先执行批处理（选择 开始质检）{Style.RESET_ALL}")
            self.pause()
            return

        print(f"  当前批次: {self.current_batch_name}")
        print(f"  录音总数: {self.current_batch.total_count}")
        print(f"  异常数: {self.current_batch.failed_count}")
        print(f"\n  将导出 Excel 表格，包含以下工作表:")
        print(f"    1. 质检总览 - 批次基本信息")
        print(f"    2. 异常清单 - 问题类型排名")
        print(f"    3. 门店汇总 - 各门店得分统计")
        print(f"    4. 详细结果 - 每条录音的完整质检结果\n")

        confirm = self.get_input("确认导出？(y/n)", "y")
        if confirm.lower() != "y":
            print("已取消")
            self.pause()
            return

        try:
            filepath = self.exporter.export_for_review(self.current_batch, self.analyzer)
            print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ 导出成功！{Style.RESET_ALL}")
            print(f"\n  文件路径: {Fore.CYAN}{filepath}{Style.RESET_ALL}")
            print(f"  请将此文件发送给质检主管进行复核")
        except Exception as e:
            print(f"\n{Fore.RED}导出失败: {e}{Style.RESET_ALL}")

        self.pause()

    # ===== 操作7: 历史记录 =====
    def action_history(self):
        filter_store = None
        filter_consultant = None
        filter_category = None
        filter_date_from = None
        filter_date_to = None

        while True:
            self.clear_screen()
            self.print_header()
            print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 历史记录 】{Style.RESET_ALL}\n")

            all_stores = self.record_manager.get_all_stores()
            all_cons = self.record_manager.get_all_consultants()

            print(f"{Style.BRIGHT}当前筛选条件:{Style.RESET_ALL}")
            filter_parts = []
            if filter_store: filter_parts.append(f"门店={filter_store}")
            if filter_consultant: filter_parts.append(f"咨询师={filter_consultant}")
            if filter_category: filter_parts.append(f"规则={RULE_CATEGORIES.get(filter_category, filter_category)}")
            if filter_date_from: filter_parts.append(f"日期>={filter_date_from}")
            if filter_date_to: filter_parts.append(f"日期<={filter_date_to}")

            if filter_parts:
                print(f"  {Fore.CYAN}{' | '.join(filter_parts)}{Style.RESET_ALL}")
            else:
                print(f"  {Fore.LIGHTBLACK_EX}(无筛选，显示全部){Style.RESET_ALL}")

            records = self.record_manager.filter_records(
                store=filter_store,
                consultant=filter_consultant,
                rule_category=filter_category,
                date_from=filter_date_from,
                date_to=filter_date_to
            )

            print(f"\n命中 {Fore.GREEN}{len(records)}{Style.RESET_ALL} 个批次\n")

            if not records:
                print(f"{Fore.LIGHTBLACK_EX}没有符合条件的批次{Style.RESET_ALL}")
            else:
                table_data = []
                for i, rec in enumerate(records[:20], 1):
                    start_time = rec["start_time"][:19].replace("T", " ")
                    cat_name = RULE_CATEGORIES.get(rec["rule_category"], rec["rule_category"])
                    pct = rec["passed_count"] / rec["total_count"] * 100 if rec["total_count"] > 0 else 0
                    table_data.append([
                        i,
                        rec["batch_name"],
                        cat_name,
                        rec["total_count"],
                        f"{pct:.1f}%",
                        start_time[:16]
                    ])

                print(tabulate(
                    table_data,
                    headers=["#", "批次名称", "规则", "总数", "通过率", "开始时间"],
                    tablefmt="simple"
                ))

                if len(records) > 20:
                    print(f"\n  ... 还有 {len(records) - 20} 个批次")

            print(f"\n{Style.BRIGHT}操作选项:{Style.RESET_ALL}")
            print(f"  [1] 设置筛选 - 门店")
            print(f"  [2] 设置筛选 - 咨询师")
            print(f"  [3] 设置筛选 - 规则类型")
            print(f"  [4] 设置筛选 - 日期范围")
            print(f"  [5] 清除所有筛选")
            print(f"  ---")
            print(f"  [6] 加载某个批次")
            print(f"  [7] 选择两个批次对比")
            print(f"  [0] 返回主菜单")

            choice = self.get_input("\n请选择操作", "0")

            if choice == "0":
                break

            elif choice == "1":
                if all_stores:
                    print(f"\n可用门店: {', '.join(all_stores)}")
                store = self.get_input("请输入门店名称（留空清除）", "")
                filter_store = store if store else None

            elif choice == "2":
                if all_cons:
                    print(f"\n可用咨询师: {', '.join(all_cons[:10])}...")
                consultant = self.get_input("请输入咨询师姓名（留空清除）", "")
                filter_consultant = consultant if consultant else None

            elif choice == "3":
                print(f"\n规则类型:")
                for i, (key, name) in enumerate(RULE_CATEGORIES.items(), 1):
                    print(f"  [{i}] {name}")
                cat_choice = self.get_input("请选择（留空清除）", "")
                if cat_choice.isdigit():
                    cats = list(RULE_CATEGORIES.keys())
                    idx = int(cat_choice) - 1
                    if 0 <= idx < len(cats):
                        filter_category = cats[idx]
                else:
                    filter_category = None

            elif choice == "4":
                date_from = self.get_input("开始日期 (YYYY-MM-DD，留空不限)", "")
                date_to = self.get_input("结束日期 (YYYY-MM-DD，留空不限)", "")
                filter_date_from = date_from if date_from else None
                filter_date_to = date_to if date_to else None

            elif choice == "5":
                filter_store = None
                filter_consultant = None
                filter_category = None
                filter_date_from = None
                filter_date_to = None
                print(f"\n{Fore.GREEN}已清除所有筛选{Style.RESET_ALL}")

            elif choice == "6":
                if not records:
                    print(f"\n{Fore.YELLOW}没有符合条件的批次{Style.RESET_ALL}")
                    self.pause()
                    continue
                idx = self.get_input("请输入批次序号", "1")
                if idx.isdigit() and 1 <= int(idx) <= len(records):
                    self._load_history(records[int(idx) - 1])
                    self.pause()

            elif choice == "7":
                if len(records) < 2:
                    print(f"\n{Fore.YELLOW}至少需要2个批次才能对比{Style.RESET_ALL}")
                    self.pause()
                    continue

                idx1 = self.get_input("请输入第一个批次序号（较旧）", "1")
                idx2 = self.get_input("请输入第二个批次序号（较新）", "2")

                if idx1.isdigit() and idx2.isdigit() \
                   and 1 <= int(idx1) <= len(records) \
                   and 1 <= int(idx2) <= len(records) \
                   and idx1 != idx2:

                    batch_a = self._load_batch_to_memory(records[int(idx1) - 1])
                    batch_b = self._load_batch_to_memory(records[int(idx2) - 1])

                    if batch_a and batch_b:
                        self._show_comparison(batch_a, batch_b)
                    else:
                        print(f"\n{Fore.RED}批次加载失败{Style.RESET_ALL}")
                else:
                    print(f"\n{Fore.YELLOW}序号无效{Style.RESET_ALL}")
                self.pause()

            else:
                print(f"\n{Fore.YELLOW}无效选项{Style.RESET_ALL}")
                self.pause()

    def _load_batch_to_memory(self, record_info: dict) -> object:
        from models import Recording, RecordingQAResult, CheckItemResult, BatchRecord

        batch_id = record_info["batch_id"]
        data = self.record_manager.load_batch(batch_id)

        if not data:
            return None

        batch = BatchRecord(
            batch_id=data["batch_id"],
            batch_name=data["batch_name"],
            start_time=datetime.fromisoformat(data["start_time"]),
            end_time=datetime.fromisoformat(data["end_time"]) if data.get("end_time") else None,
            total_count=data["total_count"],
            passed_count=data["passed_count"],
            failed_count=data["failed_count"],
            rule_category=data["rule_category"],
            deal_filter=data["deal_filter"],
            source_dir=data["source_dir"]
        )

        for r_data in data["results"]:
            rec = Recording(
                file_path=r_data["file_path"],
                file_name=r_data["file_name"],
                store=r_data["store"],
                consultant=r_data["consultant"],
                record_date=r_data["record_date"],
                record_time=r_data["record_time"],
                deal_status=r_data["deal_status"],
                duration_seconds=r_data["duration_seconds"],
                transcript="",
                category=r_data["category"]
            )

            result = RecordingQAResult(
                recording=rec,
                total_score=r_data["total_score"],
                ban_word_hits=r_data["ban_word_hits"],
                interruption_count=r_data["interruption_count"],
                price_validity_clear=r_data["price_validity_clear"],
                preop_mentioned=r_data["preop_mentioned"],
                postop_mentioned=r_data["postop_mentioned"],
                issues=r_data["issues"],
                check_results=[
                    CheckItemResult(
                        rule_id=cr["rule_id"],
                        rule_name=cr["rule_name"],
                        passed=cr["passed"],
                        score=cr["score"],
                        detail=cr["detail"],
                        evidence=cr.get("evidence", [])
                    )
                    for cr in r_data["check_results"]
                ]
            )
            batch.results.append(result)

        return batch

    def _load_history(self, record_info: dict):
        batch = self._load_batch_to_memory(record_info)
        if batch:
            self.current_batch = batch
            self.current_batch_name = record_info["batch_name"]
            self.current_category = record_info["rule_category"]
            self.current_deal_filter = record_info["deal_filter"]

            print(f"\n{Fore.GREEN}✓ 已加载批次: {record_info['batch_name']}{Style.RESET_ALL}")
            print(f"  可在【异常清单】【门店汇总】【结果导出】中查看")
        else:
            print(f"\n{Fore.RED}加载失败{Style.RESET_ALL}")

    def _show_comparison(self, batch_a, batch_b):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 批次对比 】{Style.RESET_ALL}")
        print(f"\n  {batch_a.batch_name}  →  {batch_b.batch_name}\n")

        result = self.comparator.compare_batches(batch_a, batch_b)

        a = result["batch_a"]
        b = result["batch_b"]

        print(f"{Style.BRIGHT}📊 总体指标对比:{Style.RESET_ALL}\n")

        diff_color = Fore.GREEN if b['pass_rate'] >= a['pass_rate'] else Fore.RED
        diff_sign = "+" if (b['pass_rate'] - a['pass_rate']) >= 0 else ""

        table_data = [
            ["录音总数", a['total'], b['total'],
             self._format_diff_num(b['total'] - a['total'])],
            ["通过数", a['passed'], b['passed'],
             self._format_diff_num(b['passed'] - a['passed'], good='+')],
            ["低分异常数", a['low_count'], b['low_count'],
             self._format_diff_num(b['low_count'] - a['low_count'], good='-')],
            ["通过率", f"{a['pass_rate']}%", f"{b['pass_rate']}%",
             f"{diff_color}{diff_sign}{round(b['pass_rate'] - a['pass_rate'], 1)}%{Style.RESET_ALL}"],
            ["平均分", a['avg_score'], b['avg_score'],
             self._format_diff_num(round(b['avg_score'] - a['avg_score'], 1), good='+')],
        ]

        print(tabulate(table_data, headers=["指标", a['name'], b['name'], "变化"], tablefmt="simple"))

        print(f"\n{Style.BRIGHT}🏪 门店平均分对比（按变化升序）:{Style.RESET_ALL}\n")

        store_comparison = sorted(
            result["store_comparison"],
            key=lambda x: x["avg_diff"] if x["avg_diff"] is not None else -999
        )

        store_table = []
        for sc in store_comparison:
            diff_str = self._format_diff_num(sc["avg_diff"], good='+') if sc["avg_diff"] is not None else "-"
            store_table.append([
                sc["store"],
                sc["avg_a"] if sc["avg_a"] is not None else "-",
                sc["avg_b"] if sc["avg_b"] is not None else "-",
                diff_str,
            ])

        print(tabulate(store_table, headers=["门店", f"{a['name']}(分)", f"{b['name']}(分)", "变化"], tablefmt="simple"))

        print(f"\n{Style.BRIGHT}🔝 问题类型 Top10 对比:{Style.RESET_ALL}\n")

        issue_table = []
        for i, ic in enumerate(result["issue_comparison"], 1):
            diff_str = self._format_diff_num(ic["diff"], good='-')
            issue_table.append([
                i,
                ic["issue"],
                ic["count_a"],
                ic["count_b"],
                diff_str,
            ])

        print(tabulate(issue_table, headers=["排名", "问题类型", a['name'], b['name'], "变化"], tablefmt="simple"))

        print(f"\n{Style.BRIGHT}👤 咨询师对比（高风险排前，前10）:{Style.RESET_ALL}\n")

        cons_comparison = result["consultant_comparison"][:10]
        cons_table = []
        for cc in cons_comparison:
            avg_diff_str = self._format_diff_num(cc["avg_diff"], good='+') if cc["avg_diff"] is not None else "-"
            low_diff_str = self._format_diff_num(cc["low_diff"], good='-') if cc["low_diff"] is not None else "-"

            issues_a = set(i for i, _ in cc.get("top_issues_a", []))
            issues_b = set(i for i, _ in cc.get("top_issues_b", []))
            new_issues = issues_b - issues_a
            gone_issues = issues_a - issues_b
            issue_change_parts = []
            if new_issues:
                issue_change_parts.append(f"+{','.join(list(new_issues)[:2])}")
            if gone_issues:
                issue_change_parts.append(f"-{','.join(list(gone_issues)[:2])}")
            issue_change_str = " ".join(issue_change_parts) if issue_change_parts else "—"

            cons_table.append([
                cc["store"],
                cc["consultant"],
                cc["avg_a"] if cc["avg_a"] is not None else "-",
                cc["avg_b"] if cc["avg_b"] is not None else "-",
                avg_diff_str,
                cc["low_a"],
                cc["low_b"],
                low_diff_str,
                issue_change_str,
            ])

        print(tabulate(
            cons_table,
            headers=["门店", "咨询师", f"{a['name']}(分)", f"{b['name']}(分)", "分变化",
                     f"{a['name']}(低分)", f"{b['name']}(低分)", "低分变化", "问题变化"],
            tablefmt="simple",
            maxcolwidths=[None, None, None, None, None, None, None, None, 25]
        ))

        print(f"\n{Style.BRIGHT}操作选项:{Style.RESET_ALL}")
        print(f"  [1] 导出对比表到 Excel（含咨询师对比页）")
        print(f"  [2] 将较新批次设为当前批次")
        print(f"  [0] 返回")

        choice = self.get_input("\n请选择操作", "0")
        if choice == "1":
            try:
                filepath = self.exporter.export_comparison(batch_a, batch_b, self.analyzer)
                print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ 对比表导出成功！{Style.RESET_ALL}")
                print(f"  文件路径: {Fore.CYAN}{filepath}{Style.RESET_ALL}")
            except Exception as e:
                print(f"\n{Fore.RED}导出失败: {e}{Style.RESET_ALL}")
        elif choice == "2":
            self.current_batch = batch_b
            self.current_batch_name = batch_b.batch_name
            self.current_category = batch_b.rule_category
            self.current_deal_filter = batch_b.deal_filter
            print(f"\n{Fore.GREEN}✓ 已将 [{batch_b.batch_name}] 设为当前批次{Style.RESET_ALL}")

    def _format_diff_num(self, value, good='+'):
        if value is None:
            return "-"
        if value > 0:
            color = Fore.GREEN if good == '+' else Fore.RED
            return f"{color}+{value}{Style.RESET_ALL}"
        elif value < 0:
            color = Fore.RED if good == '+' else Fore.GREEN
            return f"{color}{value}{Style.RESET_ALL}"
        else:
            return "0"

    # ===== 操作8: 周会模式 =====
    def action_weekly(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 周会模式 】{Style.RESET_ALL}")
        print(f"  选定日期范围后按周自动汇总，识别风险门店和退步咨询师\n")

        date_from = self.get_input("开始日期 (YYYY-MM-DD)", "2024-01-01")
        date_to = self.get_input("结束日期 (YYYY-MM-DD，留空到今天)", "")

        print(f"\n{Fore.CYAN}正在加载历史批次并分析...{Style.RESET_ALL}")

        result = self.weekly_analyzer.analyze_weekly(date_from, date_to)
        batches_info = result["batches"]

        if not batches_info:
            print(f"\n{Fore.YELLOW}该日期范围内没有找到批次记录{Style.RESET_ALL}")
            self.pause()
            return

        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 周会模式 - 复盘分析 】{Style.RESET_ALL}")
        print(f"  日期范围: {date_from} ~ {date_to or '今天'}")
        print(f"  涉及批次: {len(batches_info)} 个\n")

        print(f"{Style.BRIGHT}📅 批次概览:{Style.RESET_ALL}\n")
        table = [[b["name"], b["date"], b["total"], b["avg"]] for b in batches_info]
        print(tabulate(table, headers=["批次名称", "日期", "录音数", "均分"], tablefmt="simple"))

        weeks = result["weeks"]
        if weeks:
            print(f"\n{Style.BRIGHT}📊 按周趋势:{Style.RESET_ALL}\n")
            week_table = []
            for w in weeks:
                stores_str = "  ".join([f"{s}:{v}" for s, v in list(w["store_avgs"].items())[:3]])
                week_table.append([w["week"], w["batch_count"], w["total_recordings"],
                                   w["avg_score"], w["low_count"], stores_str])
            print(tabulate(week_table, headers=["周", "批次数", "录音数", "均分", "低分", "门店均分"], tablefmt="simple"))

        risk_stores = result["risk_stores"]
        if risk_stores:
            print(f"\n{Fore.RED}{Style.BRIGHT}⚠ 重点风险门店:{Style.RESET_ALL}\n")
            for rs in risk_stores[:5]:
                reasons = "; ".join(rs["reasons"])
                print(f"  {Fore.RED}• {rs['store']}{Style.RESET_ALL}  均分:{rs['avg_score']}  趋势:{rs['trend']:+.1f}")
                print(f"    原因: {reasons}")
                if rs["top_issues"]:
                    issues = "; ".join([f"{i}({c})" for i, c in rs["top_issues"]])
                    print(f"    主要问题: {Fore.LIGHTBLACK_EX}{issues}{Style.RESET_ALL}")
        else:
            print(f"\n{Fore.GREEN}✓ 无重点风险门店{Style.RESET_ALL}")

        declining = result["declining_consultants"]
        if declining:
            print(f"\n{Fore.RED}{Style.BRIGHT}📉 退步咨询师:{Style.RESET_ALL}\n")
            for dc in declining[:5]:
                print(f"  {Fore.YELLOW}• {dc['store']} - {dc['consultant']}{Style.RESET_ALL}"
                      f"  最新:{dc['latest_score']}分  趋势:{dc['trend']:+.1f}  低分:{dc['low_count']}个")
                if dc["top_issues"]:
                    issues = "; ".join([f"{i}({c})" for i, c in dc["top_issues"]])
                    print(f"    主要问题: {Fore.LIGHTBLACK_EX}{issues}{Style.RESET_ALL}")
        else:
            print(f"\n{Fore.GREEN}✓ 无退步咨询师{Style.RESET_ALL}")

        recurring = result["recurring_issues"]
        if recurring:
            print(f"\n{Fore.YELLOW}{Style.BRIGHT}🔄 反复出现的问题:{Style.RESET_ALL}\n")
            table = [[i + 1, ri["issue"], ri["frequency"], ri["total_count"], ri["avg_per_batch"]]
                     for i, ri in enumerate(recurring)]
            print(tabulate(table, headers=["#", "问题类型", "出现批次", "总次数", "批次均次"], tablefmt="simple"))
        else:
            print(f"\n{Fore.GREEN}✓ 无反复出现的问题{Style.RESET_ALL}")

        print(f"\n{Style.BRIGHT}操作选项:{Style.RESET_ALL}")
        print(f"  [1] 导出周会包（Excel含全部工作表）")
        print(f"  [0] 返回")

        choice = self.get_input("\n请选择操作", "0")
        batch_objs = result.get("batch_objects", [])
        if choice == "1" and len(batch_objs) >= 1:
            try:
                if len(batch_objs) == 1:
                    filepath = self.weekly_analyzer.export_weekly_package(
                        result, self.exporter, batch_objs[0]
                    )
                else:
                    filepath = self.weekly_analyzer.export_weekly_package(
                        result, self.exporter, batch_objs[0], batch_objs[-1]
                    )
                print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ 周会包导出成功！{Style.RESET_ALL}")
                print(f"  {Fore.CYAN}{filepath}{Style.RESET_ALL}")
                sheet_names = ["复盘摘要"]
                if len(batch_objs) >= 2:
                    sheet_names += ["对比总览", "门店对比", "问题对比", "咨询师对比"]
                else:
                    sheet_names += ["复核总览", "问题清单", "门店汇总", "明细"]
                sheet_names += ["周趋势", "风险门店", "退步咨询师", "反复问题", "建议复听清单"]
                print(f"  包含工作表: {', '.join(sheet_names)}")
            except Exception as e:
                print(f"\n{Fore.RED}导出失败: {e}{Style.RESET_ALL}")
        elif choice == "1":
            print(f"\n{Fore.YELLOW}没有批次数据可以导出{Style.RESET_ALL}")

        self.pause()

    # ===== 操作9: 咨询师下钻 =====
    def action_consultant_drill(self):
        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 咨询师下钻 】{Style.RESET_ALL}")
        print(f"  选择咨询师后查看跨批次走势、低分样本和问题变化\n")

        all_consultants = self.record_manager.get_all_consultants()
        all_stores = self.record_manager.get_all_stores()

        if not all_consultants:
            print(f"{Fore.YELLOW}暂无历史记录{Style.RESET_ALL}")
            self.pause()
            return

        print(f"{Style.BRIGHT}可用门店:{Style.RESET_ALL} {', '.join(all_stores)}")
        store = self.get_input("\n请输入门店名称", all_stores[0] if all_stores else "")

        print(f"\n{Style.BRIGHT}可用咨询师:{Style.RESET_ALL} {', '.join(all_consultants[:8])}")
        consultant = self.get_input("请输入咨询师姓名", all_consultants[0] if all_consultants else "")

        if not store or not consultant:
            print(f"\n{Fore.YELLOW}请输入门店和咨询师{Style.RESET_ALL}")
            self.pause()
            return

        print(f"\n{Fore.CYAN}正在加载 {store} - {consultant} 的历史数据...{Style.RESET_ALL}")

        all_records = self.record_manager.list_records()
        batches = []
        for rec in all_records:
            data = self.record_manager.load_batch(rec["batch_id"])
            if not data:
                continue
            has_consultant = any(
                r.get("store") == store and r.get("consultant") == consultant
                for r in data.get("results", [])
            )
            if has_consultant:
                from models import Recording, RecordingQAResult, CheckItemResult, BatchRecord
                batch = BatchRecord(
                    batch_id=data["batch_id"],
                    batch_name=data["batch_name"],
                    start_time=datetime.fromisoformat(data["start_time"]),
                    end_time=datetime.fromisoformat(data["end_time"]) if data.get("end_time") else None,
                    total_count=data["total_count"],
                    passed_count=data["passed_count"],
                    failed_count=data["failed_count"],
                    rule_category=data["rule_category"],
                    deal_filter=data["deal_filter"],
                    source_dir=data["source_dir"]
                )
                for r_data in data["results"]:
                    rec_obj = Recording(
                        file_path=r_data["file_path"],
                        file_name=r_data["file_name"],
                        store=r_data["store"],
                        consultant=r_data["consultant"],
                        record_date=r_data["record_date"],
                        record_time=r_data["record_time"],
                        deal_status=r_data["deal_status"],
                        duration_seconds=r_data["duration_seconds"],
                        transcript="",
                        category=r_data["category"]
                    )
                    result_obj = RecordingQAResult(
                        recording=rec_obj,
                        total_score=r_data["total_score"],
                        ban_word_hits=r_data["ban_word_hits"],
                        interruption_count=r_data["interruption_count"],
                        price_validity_clear=r_data["price_validity_clear"],
                        preop_mentioned=r_data["preop_mentioned"],
                        postop_mentioned=r_data["postop_mentioned"],
                        issues=r_data["issues"],
                        check_results=[
                            CheckItemResult(
                                rule_id=cr["rule_id"], rule_name=cr["rule_name"],
                                passed=cr["passed"], score=cr["score"],
                                detail=cr["detail"], evidence=cr.get("evidence", [])
                            )
                            for cr in r_data["check_results"]
                        ]
                    )
                    batch.results.append(result_obj)
                batches.append(batch)

        if not batches:
            print(f"\n{Fore.YELLOW}未找到 {store} - {consultant} 的历史数据{Style.RESET_ALL}")
            self.pause()
            return

        batches.sort(key=lambda b: b.start_time)

        trend = self.comparator.get_consultant_trend(batches, store, consultant)

        self.clear_screen()
        self.print_header()
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}【 咨询师下钻 】{Style.RESET_ALL}")
        print(f"  {store} - {consultant}  跨 {trend['batch_count']} 个批次 (按时间从早到晚)\n")

        print(f"{Style.BRIGHT}📈 分数走势 (早→晚):{Style.RESET_ALL}\n")
        trend_table = []
        for t in trend["trend"]:
            score_bar_len = int(t["avg_score"] / 5)
            score_bar = "█" * score_bar_len
            issues_str = "; ".join([f"{i}({c})" for i, c in t["top_issues"][:2]])
            trend_table.append([
                t["batch_name"],
                t["avg_score"],
                score_bar,
                t["low_count"],
                f"{t['total_count']}条",
                issues_str,
            ])

        print(tabulate(trend_table, headers=["批次", "均分", "分布", "低分", "录音", "主要问题"], tablefmt="simple"))

        print(f"\n{Style.BRIGHT}🎧 低分录音样本（按批次分组，可直接复制路径复听）:{Style.RESET_ALL}\n")
        all_low_recordings = []
        for t in trend["trend"]:
            if t["low_recordings"]:
                print(f"  {Style.BRIGHT}[{t['batch_name']}]{Style.RESET_ALL} ({len(t['low_recordings'])}个低分)")
                for lr in t["low_recordings"][:5]:
                    score_color = Fore.RED if lr["score"] < 40 else Fore.YELLOW
                    print(f"    {score_color}{lr['score']}分{Style.RESET_ALL}  {lr['file_name']}")
                    print(f"    {Fore.CYAN}{lr['file_path']}{Style.RESET_ALL}")
                    if lr["issues"]:
                        print(f"    {Fore.LIGHTBLACK_EX}问题: {'; '.join(lr['issues'][:2])}{Style.RESET_ALL}")
                    lr["batch_name"] = t["batch_name"]
                    lr["store"] = store
                    lr["consultant"] = consultant
                    all_low_recordings.append(lr)
                print()

        print(f"\n{Style.BRIGHT}🔄 高频问题变化 (按时间顺序):{Style.RESET_ALL}\n")
        all_issues_timeline = defaultdict(list)
        for t in trend["trend"]:
            batch_issue_counts = dict(t["top_issues"])
            for issue in all_issues_timeline.keys():
                all_issues_timeline[issue].append(batch_issue_counts.get(issue, 0))
            for issue, count in t["top_issues"]:
                if issue not in all_issues_timeline:
                    all_issues_timeline[issue] = [0] * (len(all_issues_timeline[list(all_issues_timeline.keys())[0]]) if all_issues_timeline else 0) + [count]

        issue_table = []
        for issue, counts in sorted(all_issues_timeline.items(), key=lambda x: sum(x[1]), reverse=True):
            trend_str = " → ".join(str(c) for c in counts)
            if len(counts) >= 2:
                direction = "↑" if counts[-1] > counts[0] else ("↓" if counts[-1] < counts[0] else "—")
            else:
                direction = "—"
            issue_table.append([issue, trend_str, direction])

        print(tabulate(issue_table, headers=["问题类型", "各批次次数", "趋势"], tablefmt="simple"))

        if all_low_recordings:
            print(f"\n{Style.BRIGHT}📋 复听任务清单（按问题类型分组）:{Style.RESET_ALL}\n")
            by_issue = defaultdict(list)
            for lr in all_low_recordings:
                for issue in lr["issues"]:
                    by_issue[issue].append(lr)

            for issue, records in sorted(by_issue.items(), key=lambda x: len(x[1]), reverse=True):
                print(f"  {Style.BRIGHT}【{issue}】{Style.RESET_ALL} ({len(records)}条)")
                for lr in records[:3]:
                    print(f"    {Fore.YELLOW}{lr['score']}分{Style.RESET_ALL} [{lr['batch_name']}] {lr['file_name']}")
                if len(records) > 3:
                    print(f"    ... 还有 {len(records) - 3} 条")
                print()

            print(f"\n{Style.BRIGHT}操作选项:{Style.RESET_ALL}")
            print(f"  [1] 导出复听任务清单（按问题类型/批次/分数分组）")
            print(f"  [0] 返回")

            choice = self.get_input("\n请选择操作", "0")
            if choice == "1":
                filepath = self._export_review_task_list(store, consultant, all_low_recordings, batches)
                print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ 复听任务清单导出成功！{Style.RESET_ALL}")
                print(f"  {Fore.CYAN}{filepath}{Style.RESET_ALL}")
                self.pause()
            return

        self.pause()

    # ===== 主循环 =====
    def run(self):
        while True:
            self.clear_screen()
            self.print_header()
            self.print_status_bar()
            self.print_menu()

            choice = self.get_input("\n请选择操作", "")

            if choice == "1":
                self.action_import()
            elif choice == "2":
                self.action_rules()
            elif choice == "3":
                self.action_batch()
            elif choice == "4":
                self.action_issues()
            elif choice == "5":
                self.action_stores()
            elif choice == "6":
                self.action_export()
            elif choice == "7":
                self.action_history()
            elif choice == "8":
                self.action_weekly()
            elif choice == "9":
                self.action_consultant_drill()
            elif choice == "0":
                print(f"\n{Fore.CYAN}感谢使用，再见！{Style.RESET_ALL}\n")
                break
            else:
                print(f"\n{Fore.RED}无效选项，请重新选择{Style.RESET_ALL}")
                self.pause()


def main():
    try:
        cli = QACLI()
        cli.run()
    except KeyboardInterrupt:
        print(f"\n\n{Fore.CYAN}感谢使用，再见！{Style.RESET_ALL}\n")
    except Exception as e:
        print(f"\n{Fore.RED}程序异常: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
        input("按回车键退出...")


if __name__ == "__main__":
    main()
