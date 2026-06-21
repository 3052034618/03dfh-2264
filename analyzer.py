import os
import json
from collections import defaultdict
from typing import List, Dict
from datetime import datetime
from models import BatchRecord, RecordingQAResult, StoreSummary
from config import EXPORTS_DIR, RECORDS_DIR, RULE_CATEGORIES, DEAL_STATUS, LOW_SCORE_THRESHOLD
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ResultAnalyzer:
    def __init__(self):
        pass

    def get_store_summaries(self, batch: BatchRecord) -> List[StoreSummary]:
        store_data = defaultdict(list)
        for result in batch.results:
            store_data[result.recording.store].append(result)

        summaries = []
        for store_name, results in store_data.items():
            total = len(results)
            avg_score = sum(r.total_score for r in results) / total if total > 0 else 0
            pass_count = sum(1 for r in results if not r.is_low_score)
            pass_rate = pass_count / total if total > 0 else 0
            issue_count = sum(len(r.issues) for r in results)

            issue_counts = defaultdict(int)
            for r in results:
                for issue in r.issues:
                    issue_counts[issue] += 1

            top_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)[:5]

            summaries.append(StoreSummary(
                store_name=store_name,
                total_count=total,
                avg_score=round(avg_score, 1),
                pass_rate=round(pass_rate * 100, 1),
                issue_count=issue_count,
                top_issues=top_issues
            ))

        summaries.sort(key=lambda x: x.avg_score)
        return summaries

    def get_consultant_summaries(self, batch: BatchRecord) -> List[dict]:
        consultant_data = defaultdict(list)
        for result in batch.results:
            key = f"{result.recording.store}|{result.recording.consultant}"
            consultant_data[key].append(result)

        summaries = []
        for key, results in consultant_data.items():
            store, consultant = key.split("|", 1)
            total = len(results)
            avg_score = sum(r.total_score for r in results) / total if total > 0 else 0
            low_count = sum(1 for r in results if r.is_low_score)
            pass_rate = (total - low_count) / total if total > 0 else 0
            issue_count = sum(len(r.issues) for r in results)

            issue_counts = defaultdict(int)
            for r in results:
                for issue in r.issues:
                    issue_counts[issue] += 1

            top_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)[:3]

            summaries.append({
                "store": store,
                "consultant": consultant,
                "total_count": total,
                "avg_score": round(avg_score, 1),
                "low_count": low_count,
                "pass_rate": round(pass_rate * 100, 1),
                "issue_count": issue_count,
                "top_issues": top_issues
            })

        summaries.sort(key=lambda x: x["avg_score"])
        return summaries

    def get_top_issues(self, batch: BatchRecord, top_n: int = 10) -> List[tuple]:
        issue_counts = defaultdict(int)
        for result in batch.results:
            for issue in result.issues:
                issue_counts[issue] += 1

        sorted_issues = sorted(issue_counts.items(), key=lambda x: x[1], reverse=True)
        return sorted_issues[:top_n]

    def get_low_score_list(self, batch: BatchRecord) -> List[RecordingQAResult]:
        low_scores = [r for r in batch.results if r.is_low_score]
        low_scores.sort(key=lambda x: x.total_score)
        return low_scores

    def get_category_stats(self, batch: BatchRecord) -> dict:
        total = len(batch.results)
        passed = batch.passed_count
        failed = batch.failed_count
        avg_score = sum(r.total_score for r in batch.results) / total if total > 0 else 0

        category_stats = defaultdict(lambda: {"count": 0, "score_sum": 0})

        for r in batch.results:
            for check in r.check_results:
                category_stats[check.rule_name]["count"] += 1
                category_stats[check.rule_name]["score_sum"] += check.score

        stats = []
        for name, data in category_stats.items():
            avg = data["score_sum"] / data["count"] if data["count"] > 0 else 0
            stats.append({
                "name": name,
                "avg_score": round(avg, 1),
                "count": data["count"]
            })

        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": round(passed / total * 100, 1) if total > 0 else 0,
            "avg_score": round(avg_score, 1),
            "category_stats": stats
        }


class RecordManager:
    def __init__(self):
        self.records_dir = RECORDS_DIR
        os.makedirs(self.records_dir, exist_ok=True)

    def save_batch(self, batch: BatchRecord) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"batch_{batch.batch_id}_{timestamp}.json"
        filepath = os.path.join(self.records_dir, filename)

        data = {
            "batch_id": batch.batch_id,
            "batch_name": batch.batch_name,
            "start_time": batch.start_time.isoformat(),
            "end_time": batch.end_time.isoformat() if batch.end_time else None,
            "total_count": batch.total_count,
            "passed_count": batch.passed_count,
            "failed_count": batch.failed_count,
            "rule_category": batch.rule_category,
            "deal_filter": batch.deal_filter,
            "source_dir": batch.source_dir,
            "results": [self._serialize_result(r) for r in batch.results]
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return filepath

    def _serialize_result(self, result: RecordingQAResult) -> dict:
        return {
            "file_path": result.recording.file_path,
            "file_name": result.recording.file_name,
            "store": result.recording.store,
            "consultant": result.recording.consultant,
            "record_date": result.recording.record_date,
            "record_time": result.recording.record_time,
            "deal_status": result.recording.deal_status,
            "duration_seconds": result.recording.duration_seconds,
            "category": result.recording.category,
            "total_score": result.total_score,
            "ban_word_hits": result.ban_word_hits,
            "interruption_count": result.interruption_count,
            "price_validity_clear": result.price_validity_clear,
            "preop_mentioned": result.preop_mentioned,
            "postop_mentioned": result.postop_mentioned,
            "issues": result.issues,
            "check_results": [
                {
                    "rule_id": cr.rule_id,
                    "rule_name": cr.rule_name,
                    "passed": cr.passed,
                    "score": cr.score,
                    "detail": cr.detail,
                    "evidence": cr.evidence
                }
                for cr in result.check_results
            ]
        }

    def _list_all_raw(self) -> List[dict]:
        records = []
        for filename in os.listdir(self.records_dir):
            if not filename.endswith(".json"):
                continue
            filepath = os.path.join(self.records_dir, filename)
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                stores = set()
                consultants = set()
                for r in data.get("results", []):
                    stores.add(r.get("store", ""))
                    consultants.add(r.get("consultant", ""))

                records.append({
                    "batch_id": data["batch_id"],
                    "batch_name": data["batch_name"],
                    "start_time": data["start_time"],
                    "end_time": data.get("end_time"),
                    "total_count": data["total_count"],
                    "passed_count": data["passed_count"],
                    "failed_count": data["failed_count"],
                    "rule_category": data["rule_category"],
                    "deal_filter": data["deal_filter"],
                    "stores": list(stores),
                    "consultants": list(consultants),
                    "filepath": filepath
                })
            except Exception:
                continue

        records.sort(key=lambda x: x["start_time"], reverse=True)
        return records

    def list_records(self) -> List[dict]:
        return self._list_all_raw()

    def filter_records(self, store: str = None, consultant: str = None,
                       rule_category: str = None, date_from: str = None,
                       date_to: str = None) -> List[dict]:
        records = self._list_all_raw()
        filtered = []

        for rec in records:
            if store and store not in rec["stores"]:
                continue
            if consultant and consultant not in rec["consultants"]:
                continue
            if rule_category and rec["rule_category"] != rule_category:
                continue
            if date_from:
                rec_date = rec["start_time"][:10]
                if rec_date < date_from:
                    continue
            if date_to:
                rec_date = rec["start_time"][:10]
                if rec_date > date_to:
                    continue
            filtered.append(rec)

        return filtered

    def get_all_stores(self) -> List[str]:
        records = self._list_all_raw()
        stores = set()
        for rec in records:
            for s in rec["stores"]:
                if s:
                    stores.add(s)
        return sorted(list(stores))

    def get_all_consultants(self) -> List[str]:
        records = self._list_all_raw()
        consultants = set()
        for rec in records:
            for c in rec["consultants"]:
                if c:
                    consultants.add(c)
        return sorted(list(consultants))

    def load_batch(self, batch_id: str) -> dict:
        for filename in os.listdir(self.records_dir):
            if not filename.endswith(".json"):
                continue
            if batch_id in filename:
                filepath = os.path.join(self.records_dir, filename)
                with open(filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
        return {}


class ExcelExporter:
    def __init__(self):
        self.exports_dir = EXPORTS_DIR
        os.makedirs(self.exports_dir, exist_ok=True)

    def export_for_review(self, batch: BatchRecord, analyzer: ResultAnalyzer) -> str:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "质检总览"
        self._fill_overview_sheet(ws1, batch, analyzer)

        ws2 = wb.create_sheet("异常清单")
        self._fill_issues_sheet(ws2, batch, analyzer)

        ws3 = wb.create_sheet("门店汇总")
        self._fill_stores_sheet(ws3, batch, analyzer)

        ws4 = wb.create_sheet("详细结果")
        self._fill_detail_sheet(ws4, batch)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"质检报告_{batch.batch_name}_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)

        wb.save(filepath)
        return filepath

    def _set_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _set_cell_style(self, cell):
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _fill_overview_sheet(self, ws, batch, analyzer):
        stats = analyzer.get_category_stats(batch)

        ws.merge_cells("A1:D1")
        ws["A1"] = f"质检报告 - " + batch.batch_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        info = [
            ["批次ID", batch.batch_id],
            ["质检规则", RULE_CATEGORIES.get(batch.rule_category, batch.rule_category)],
            ["成交筛选", DEAL_STATUS.get(batch.deal_filter, batch.deal_filter)],
            ["源文件夹", batch.source_dir],
            ["开始时间", batch.start_time.strftime("%Y-%m-%d %H:%M:%S")],
            ["结束时间", batch.end_time.strftime("%Y-%m-%d %H:%M:%S") if batch.end_time else ""],
            ["录音总数", stats["total"]],
            ["通过数量", stats["passed"]],
            ["异常数量", stats["failed"]],
            ["通过率", f"{stats['pass_rate']}%"],
            ["平均分", stats["avg_score"]],
        ]

        row = 3
        for key, value in info:
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40

    def _fill_issues_sheet(self, ws, batch, analyzer):
        ws.merge_cells("A1:H1")
        title_cell = ws.cell(row=1, column=1, value="⚠ 低分录音清单（按得分升序）")
        title_cell.font = Font(bold=True, size=14, color="C00000")
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        headers = ["排名", "门店", "咨询师", "得分", "问题数量", "主要问题", "文件路径", "转写来源"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._set_header_style(cell)

        low_scores = analyzer.get_low_score_list(batch)
        for row_idx, result in enumerate(low_scores, 3):
            rec = result.recording
            source = "真实转写" if getattr(rec, "transcript_source", "simulated") == "real" else "模拟转写"
            issues_str = "; ".join(result.issues[:3])
            if len(result.issues) > 3:
                issues_str += f" ...(+{len(result.issues) - 3})"

            ws.cell(row=row_idx, column=1, value=row_idx - 2)
            ws.cell(row=row_idx, column=2, value=rec.store)
            ws.cell(row=row_idx, column=3, value=rec.consultant)
            ws.cell(row=row_idx, column=4, value=result.total_score)
            ws.cell(row=row_idx, column=5, value=len(result.issues))
            ws.cell(row=row_idx, column=6, value=issues_str)
            ws.cell(row=row_idx, column=7, value=rec.file_path)
            ws.cell(row=row_idx, column=8, value=source)

            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                self._set_cell_style(cell)
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        start_row = len(low_scores) + 5
        ws.cell(row=start_row, column=1, value="问题类型统计").font = Font(bold=True, size=12)
        start_row += 1

        stat_headers = ["排名", "问题类型", "出现次数", "占比"]
        for col, header in enumerate(stat_headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            self._set_header_style(cell)

        top_issues = analyzer.get_top_issues(batch, top_n=20)
        total = batch.total_count
        for i, (issue, count) in enumerate(top_issues, 1):
            row = start_row + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=issue)
            ws.cell(row=row, column=3, value=count)
            ws.cell(row=row, column=4, value=f"{count/total*100:.1f}%")
            for col in range(1, 5):
                self._set_cell_style(ws.cell(row=row, column=col))

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 35
        ws.column_dimensions["G"].width = 60
        ws.column_dimensions["H"].width = 10

    def _fill_stores_sheet(self, ws, batch, analyzer):
        summaries = analyzer.get_store_summaries(batch)

        headers = ["门店名称", "录音数", "平均分", "通过率", "问题总数", "主要问题"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._set_header_style(cell)

        for row_idx, summary in enumerate(summaries, 2):
            ws.cell(row=row_idx, column=1, value=summary.store_name)
            ws.cell(row=row_idx, column=2, value=summary.total_count)
            ws.cell(row=row_idx, column=3, value=summary.avg_score)
            ws.cell(row=row_idx, column=4, value=f"{summary.pass_rate}%")
            ws.cell(row=row_idx, column=5, value=summary.issue_count)

            top_issues_str = "; ".join([f"{issue}({count})" for issue, count in summary.top_issues[:3]])
            ws.cell(row=row_idx, column=6, value=top_issues_str)

            if summary.avg_score < 60:
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

            for col in range(1, 7):
                self._set_cell_style(ws.cell(row=row_idx, column=col))

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 40

    def _fill_detail_sheet(self, ws, batch):
        headers = [
            "门店", "咨询师", "日期", "成交状态",
            "总得分", "禁用词", "打断次数",
            "有效期说明", "术前检查", "术后护理",
            "问题列表", "文件路径"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._set_header_style(cell)

        results = sorted(batch.results, key=lambda x: x.total_score)

        for row_idx, result in enumerate(results, 2):
            rec = result.recording
            ws.cell(row=row_idx, column=1, value=rec.store)
            ws.cell(row=row_idx, column=2, value=rec.consultant)
            ws.cell(row=row_idx, column=3, value=rec.record_date)
            ws.cell(row=row_idx, column=4, value="已成交" if rec.is_dealt else "未成交")
            ws.cell(row=row_idx, column=5, value=result.total_score)
            ws.cell(row=row_idx, column=6, value="、".join(result.ban_word_hits) if result.ban_word_hits else "无")
            ws.cell(row=row_idx, column=7, value=result.interruption_count)
            ws.cell(row=row_idx, column=8, value="是" if result.price_validity_clear else "否")
            ws.cell(row=row_idx, column=9, value="是" if result.preop_mentioned else "否")
            ws.cell(row=row_idx, column=10, value="是" if result.postop_mentioned else "否")
            ws.cell(row=row_idx, column=11, value="; ".join(result.issues))
            ws.cell(row=row_idx, column=12, value=rec.file_path)

            if result.is_low_score:
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

            for col in range(1, 13):
                self._set_cell_style(ws.cell(row=row_idx, column=col))

        widths = [12, 12, 12, 10, 8, 20, 10, 10, 10, 10, 30, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    def export_comparison(self, batch_a: BatchRecord, batch_b: BatchRecord,
                          analyzer: ResultAnalyzer) -> str:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "对比总览"
        self._fill_compare_overview_sheet(ws1, batch_a, batch_b, analyzer)

        ws2 = wb.create_sheet("门店对比")
        self._fill_compare_stores_sheet(ws2, batch_a, batch_b, analyzer)

        ws3 = wb.create_sheet("问题对比")
        self._fill_compare_issues_sheet(ws3, batch_a, batch_b, analyzer)

        ws4 = wb.create_sheet("咨询师对比")
        self._fill_compare_consultants_sheet(ws4, batch_a, batch_b, analyzer)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"批次对比_{batch_a.batch_name}_vs_{batch_b.batch_name}_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)

        wb.save(filepath)
        return filepath

    def _fill_compare_overview_sheet(self, ws, batch_a, batch_b, analyzer):
        ws.merge_cells("A1:D1")
        ws["A1"] = "批次对比报告"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        stats_a = analyzer.get_category_stats(batch_a)
        stats_b = analyzer.get_category_stats(batch_b)

        headers = ["指标", batch_a.batch_name, batch_b.batch_name, "变化"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._set_header_style(cell)

        rows = [
            ["批次ID", batch_a.batch_id, batch_b.batch_id, ""],
            ["质检规则", RULE_CATEGORIES.get(batch_a.rule_category, batch_a.rule_category),
             RULE_CATEGORIES.get(batch_b.rule_category, batch_b.rule_category), ""],
            ["录音总数", stats_a["total"], stats_b["total"],
             self._format_diff(stats_b["total"] - stats_a["total"])],
            ["通过数", stats_a["passed"], stats_b["passed"],
             self._format_diff(stats_b["passed"] - stats_a["passed"])],
            ["异常数", stats_a["failed"], stats_b["failed"],
             self._format_diff(stats_b["failed"] - stats_a["failed"], reverse=True)],
            ["通过率", f"{stats_a['pass_rate']}%", f"{stats_b['pass_rate']}%",
             self._format_pct_diff(stats_b['pass_rate'] - stats_a['pass_rate'])],
            ["平均分", stats_a["avg_score"], stats_b["avg_score"],
             self._format_diff(round(stats_b["avg_score"] - stats_a["avg_score"], 1))],
        ]

        for row_idx, row_data in enumerate(rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._set_cell_style(cell)
                if col_idx == 4 and value and value not in ["", "-"]:
                    if "↓" in str(value):
                        cell.font = Font(color="00B050", bold=True)
                    elif "↑" in str(value):
                        cell.font = Font(color="FF0000", bold=True)

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 15

    def _fill_compare_stores_sheet(self, ws, batch_a, batch_b, analyzer):
        ws.merge_cells("A1:G1")
        ws["A1"] = "门店平均分对比"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        stores_a = {s.store_name: s for s in analyzer.get_store_summaries(batch_a)}
        stores_b = {s.store_name: s for s in analyzer.get_store_summaries(batch_b)}

        all_stores = sorted(set(list(stores_a.keys()) + list(stores_b.keys())))

        headers = ["门店", f"{batch_a.batch_name}(平均分)", f"{batch_b.batch_name}(平均分)",
                   "分数变化", f"{batch_a.batch_name}(通过率)", f"{batch_b.batch_name}(通过率)", "通过率变化"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._set_header_style(cell)

        for row_idx, store in enumerate(all_stores, 3):
            sa = stores_a.get(store)
            sb = stores_b.get(store)

            avg_a = sa.avg_score if sa else 0
            avg_b = sb.avg_score if sb else 0
            rate_a = sa.pass_rate if sa else 0
            rate_b = sb.pass_rate if sb else 0

            avg_diff = round(avg_b - avg_a, 1)
            rate_diff = round(rate_b - rate_a, 1)

            row_data = [
                store,
                avg_a if sa else "-",
                avg_b if sb else "-",
                self._format_diff(avg_diff) if sa and sb else "-",
                f"{rate_a}%" if sa else "-",
                f"{rate_b}%" if sb else "-",
                self._format_pct_diff(rate_diff) if sa and sb else "-",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._set_cell_style(cell)
                if col_idx in [4, 7] and value not in ["-", ""]:
                    if "↓" in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                    elif "↑" in str(value):
                        cell.font = Font(color="00B050", bold=True)

        widths = [15, 22, 22, 12, 22, 22, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    def _fill_compare_issues_sheet(self, ws, batch_a, batch_b, analyzer):
        ws.merge_cells("A1:F1")
        ws["A1"] = "问题类型 Top10 对比"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        issues_a = dict(analyzer.get_top_issues(batch_a, top_n=20))
        issues_b = dict(analyzer.get_top_issues(batch_b, top_n=20))

        all_issues = set(list(issues_a.keys()) + list(issues_b.keys()))
        sorted_issues = sorted(all_issues, key=lambda x: issues_b.get(x, 0), reverse=True)[:15]

        headers = ["排名", "问题类型", f"{batch_a.batch_name}(次)",
                   f"{batch_b.batch_name}(次)", "变化", "趋势"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._set_header_style(cell)

        for row_idx, issue in enumerate(sorted_issues, 3):
            count_a = issues_a.get(issue, 0)
            count_b = issues_b.get(issue, 0)
            diff = count_b - count_a

            if diff > 0:
                trend = "↑ 增加"
            elif diff < 0:
                trend = "↓ 减少"
            else:
                trend = "— 持平"

            row_data = [
                row_idx - 2,
                issue,
                count_a,
                count_b,
                self._format_diff(diff, reverse=True),
                trend,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._set_cell_style(cell)
                if col_idx == 5 and value not in ["-", ""]:
                    if "↑" in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                    elif "↓" in str(value):
                        cell.font = Font(color="00B050", bold=True)
                if col_idx == 6:
                    if "↑" in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                    elif "↓" in str(value):
                        cell.font = Font(color="00B050", bold=True)

        widths = [8, 30, 18, 18, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    def _format_diff(self, value, reverse=False):
        if value > 0:
            symbol = "↑" if reverse else "↑"
            return f"+{value} {symbol}"
        elif value < 0:
            symbol = "↓" if reverse else "↓"
            return f"{value} {symbol}"
        else:
            return "0"

    def _format_pct_diff(self, value):
        if value > 0:
            return f"+{value}% ↑"
        elif value < 0:
            return f"{value}% ↓"
        else:
            return "0%"

    def _fill_compare_consultants_sheet(self, ws, batch_a, batch_b, analyzer):
        ws.merge_cells("A1:H1")
        ws["A1"] = "咨询师对比（按平均分升序）"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        cons_a = {f"{c['store']}|{c['consultant']}": c
                  for c in analyzer.get_consultant_summaries(batch_a)}
        cons_b = {f"{c['store']}|{c['consultant']}": c
                  for c in analyzer.get_consultant_summaries(batch_b)}
        all_cons = sorted(set(list(cons_a.keys()) + list(cons_b.keys())))

        sorted_cons = sorted(
            all_cons,
            key=lambda k: cons_b.get(k, {}).get("avg_score", 0)
            if cons_b.get(k) else cons_a.get(k, {}).get("avg_score", 0)
        )

        headers = ["排名", "门店", "咨询师", f"{batch_a.batch_name}(均分)",
                   f"{batch_b.batch_name}(均分)", "分数变化", f"{batch_a.batch_name}(低分)",
                   f"{batch_b.batch_name}(低分)", "低分变化", "主要问题变化"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._set_header_style(cell)

        for row_idx, key in enumerate(sorted_cons, 3):
            store, consultant = key.split("|", 1)
            ca = cons_a.get(key)
            cb = cons_b.get(key)

            avg_a = ca["avg_score"] if ca else 0
            avg_b = cb["avg_score"] if cb else 0
            low_a = ca["low_count"] if ca else 0
            low_b = cb["low_count"] if cb else 0
            avg_diff = round(avg_b - avg_a, 1) if ca and cb else None
            low_diff = low_b - low_a if ca and cb else None

            top_a = "; ".join([f"{i}({c})" for i, c in ca["top_issues"][:2]]) if ca else "-"
            top_b = "; ".join([f"{i}({c})" for i, c in cb["top_issues"][:2]]) if cb else "-"
            issue_change = f"{top_a} → {top_b}" if ca and cb else (top_b if cb else top_a)

            row_data = [
                row_idx - 2,
                store,
                consultant,
                avg_a if ca else "-",
                avg_b if cb else "-",
                self._format_diff(avg_diff) if avg_diff is not None else "-",
                low_a if ca else "-",
                low_b if cb else "-",
                self._format_diff(low_diff, reverse=True) if low_diff is not None else "-",
                issue_change,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._set_cell_style(cell)
                if col_idx in [6, 9] and value not in ["-", ""]:
                    if "↓" in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                    elif "↑" in str(value):
                        cell.font = Font(color="00B050", bold=True)

        widths = [6, 12, 12, 18, 18, 12, 16, 16, 12, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w


class BatchComparator:
    def __init__(self, analyzer: ResultAnalyzer):
        self.analyzer = analyzer

    def compare_batches(self, batch_a: BatchRecord, batch_b: BatchRecord) -> dict:
        stats_a = self.analyzer.get_category_stats(batch_a)
        stats_b = self.analyzer.get_category_stats(batch_b)

        stores_a = {s.store_name: s for s in self.analyzer.get_store_summaries(batch_a)}
        stores_b = {s.store_name: s for s in self.analyzer.get_store_summaries(batch_b)}
        all_stores = sorted(set(list(stores_a.keys()) + list(stores_b.keys())))

        store_comparison = []
        for store in all_stores:
            sa = stores_a.get(store)
            sb = stores_b.get(store)
            store_comparison.append({
                "store": store,
                "avg_a": sa.avg_score if sa else None,
                "avg_b": sb.avg_score if sb else None,
                "avg_diff": round(sb.avg_score - sa.avg_score, 1) if sa and sb else None,
                "rate_a": sa.pass_rate if sa else None,
                "rate_b": sb.pass_rate if sb else None,
                "rate_diff": round(sb.pass_rate - sa.pass_rate, 1) if sa and sb else None,
                "count_a": sa.total_count if sa else 0,
                "count_b": sb.total_count if sb else 0,
            })

        issues_a = dict(self.analyzer.get_top_issues(batch_a, top_n=20))
        issues_b = dict(self.analyzer.get_top_issues(batch_b, top_n=20))
        all_issues = sorted(
            set(list(issues_a.keys()) + list(issues_b.keys())),
            key=lambda x: issues_b.get(x, 0),
            reverse=True
        )[:10]

        issue_comparison = []
        for issue in all_issues:
            count_a = issues_a.get(issue, 0)
            count_b = issues_b.get(issue, 0)
            issue_comparison.append({
                "issue": issue,
                "count_a": count_a,
                "count_b": count_b,
                "diff": count_b - count_a,
            })

        low_a = len(self.analyzer.get_low_score_list(batch_a))
        low_b = len(self.analyzer.get_low_score_list(batch_b))

        cons_a = {f"{c['store']}|{c['consultant']}": c
                  for c in self.analyzer.get_consultant_summaries(batch_a)}
        cons_b = {f"{c['store']}|{c['consultant']}": c
                  for c in self.analyzer.get_consultant_summaries(batch_b)}
        all_cons = sorted(set(list(cons_a.keys()) + list(cons_b.keys())))

        consultant_comparison = []
        for key in all_cons:
            store, consultant = key.split("|", 1)
            ca = cons_a.get(key)
            cb = cons_b.get(key)
            consultant_comparison.append({
                "store": store,
                "consultant": consultant,
                "avg_a": ca["avg_score"] if ca else None,
                "avg_b": cb["avg_score"] if cb else None,
                "avg_diff": round(cb["avg_score"] - ca["avg_score"], 1) if ca and cb else None,
                "low_a": ca["low_count"] if ca else 0,
                "low_b": cb["low_count"] if cb else 0,
                "low_diff": cb["low_count"] - ca["low_count"] if ca and cb else None,
                "count_a": ca["total_count"] if ca else 0,
                "count_b": cb["total_count"] if cb else 0,
                "top_issues_a": ca["top_issues"] if ca else [],
                "top_issues_b": cb["top_issues"] if cb else [],
            })

        consultant_comparison.sort(key=lambda x: self._consultant_risk_score(x), reverse=True)

        return {
            "batch_a": {
                "name": batch_a.batch_name,
                "id": batch_a.batch_id,
                "total": stats_a["total"],
                "passed": stats_a["passed"],
                "failed": stats_a["failed"],
                "pass_rate": stats_a["pass_rate"],
                "avg_score": stats_a["avg_score"],
                "low_count": low_a,
            },
            "batch_b": {
                "name": batch_b.batch_name,
                "id": batch_b.batch_id,
                "total": stats_b["total"],
                "passed": stats_b["passed"],
                "failed": stats_b["failed"],
                "pass_rate": stats_b["pass_rate"],
                "avg_score": stats_b["avg_score"],
                "low_count": low_b,
            },
            "store_comparison": store_comparison,
            "issue_comparison": issue_comparison,
            "consultant_comparison": consultant_comparison,
            "low_diff": low_b - low_a,
        }

    def _consultant_risk_score(self, cc: dict) -> float:
        risk = 0.0
        avg_diff = cc.get("avg_diff")
        low_diff = cc.get("low_diff")
        if avg_diff is not None:
            risk += max(0, -avg_diff) * 2
        if low_diff is not None:
            risk += max(0, low_diff) * 3
        issues_a = set(i for i, _ in cc.get("top_issues_a", []))
        issues_b = set(i for i, _ in cc.get("top_issues_b", []))
        new_issues = issues_b - issues_a
        risk += len(new_issues) * 2
        if avg_diff is None:
            risk -= 5
        if low_diff is None:
            risk -= 3
        return risk

    def get_consultant_trend(self, batches: list, store: str, consultant: str) -> dict:
        trend_data = []
        for batch in batches:
            cons_summaries = self.analyzer.get_consultant_summaries(batch)
            matched = [c for c in cons_summaries
                       if c["store"] == store and c["consultant"] == consultant]
            if matched:
                c = matched[0]
                low_recordings = []
                for r in batch.results:
                    if r.recording.store == store and r.recording.consultant == consultant:
                        if r.is_low_score:
                            low_recordings.append({
                                "score": r.total_score,
                                "file_path": r.recording.file_path,
                                "file_name": r.recording.file_name,
                                "issues": r.issues,
                            })
                trend_data.append({
                    "batch_name": batch.batch_name,
                    "batch_id": batch.batch_id,
                    "start_time": batch.start_time.isoformat() if batch.start_time else "",
                    "avg_score": c["avg_score"],
                    "low_count": c["low_count"],
                    "total_count": c["total_count"],
                    "top_issues": c["top_issues"],
                    "low_recordings": low_recordings,
                })
        return {
            "store": store,
            "consultant": consultant,
            "batch_count": len(trend_data),
            "trend": trend_data,
        }


class WeeklyAnalyzer:
    def __init__(self, record_manager: RecordManager, analyzer: ResultAnalyzer):
        self.record_manager = record_manager
        self.analyzer = analyzer
        self.low_score_threshold = LOW_SCORE_THRESHOLD

    def _load_batches_in_range(self, date_from: str, date_to: str) -> list:
        records = self.record_manager.filter_records(
            date_from=date_from, date_to=date_to
        )
        batches = []
        for rec in records:
            data = self.record_manager.load_batch(rec["batch_id"])
            if not data:
                continue
            batch = self._data_to_batch(data)
            if batch:
                batches.append(batch)
        return batches

    def _data_to_batch(self, data: dict) -> object:
        from models import Recording, RecordingQAResult, CheckItemResult, BatchRecord
        batch = BatchRecord(
            batch_id=data["batch_id"],
            batch_name=data["batch_name"],
            start_time=datetime.fromisoformat(data["start_time"]),
            end_time=datetime.fromisoformat(data["end_time"]) if data.get("end_time") else None,
            total_count=data["total_count"],
            passed_count=data["passed_count"],
            failed_count=data["failed_count"],
            rule_category=data["rule_category"],
            deal_filter=data["deal_filter"],
            source_dir=data["source_dir"]
        )
        for r_data in data["results"]:
            rec = Recording(
                file_path=r_data["file_path"],
                file_name=r_data["file_name"],
                store=r_data["store"],
                consultant=r_data["consultant"],
                record_date=r_data["record_date"],
                record_time=r_data["record_time"],
                deal_status=r_data["deal_status"],
                duration_seconds=r_data["duration_seconds"],
                transcript="",
                category=r_data["category"]
            )
            result = RecordingQAResult(
                recording=rec,
                total_score=r_data["total_score"],
                ban_word_hits=r_data["ban_word_hits"],
                interruption_count=r_data["interruption_count"],
                price_validity_clear=r_data["price_validity_clear"],
                preop_mentioned=r_data["preop_mentioned"],
                postop_mentioned=r_data["postop_mentioned"],
                issues=r_data["issues"],
                check_results=[
                    CheckItemResult(
                        rule_id=cr["rule_id"],
                        rule_name=cr["rule_name"],
                        passed=cr["passed"],
                        score=cr["score"],
                        detail=cr["detail"],
                        evidence=cr.get("evidence", [])
                    )
                    for cr in r_data["check_results"]
                ]
            )
            batch.results.append(result)
        return batch

    def analyze_weekly(self, date_from: str, date_to: str) -> dict:
        batches = self._load_batches_in_range(date_from, date_to)
        if not batches:
            return {"batches": [], "weeks": [], "risk_stores": [],
                    "declining_consultants": [], "recurring_issues": []}

        weeks = self._group_by_week(batches)
        risk_stores = self._find_risk_stores(batches)
        declining_consultants = self._find_declining_consultants(batches)
        recurring_issues = self._find_recurring_issues(batches)

        return {
            "batches": [{"name": b.batch_name, "id": b.batch_id,
                          "date": b.start_time.strftime("%Y-%m-%d"),
                          "total": b.total_count, "avg": round(
                              sum(r.total_score for r in b.results) / b.total_count, 1
                          ) if b.total_count else 0}
                         for b in batches],
            "weeks": weeks,
            "risk_stores": risk_stores,
            "declining_consultants": declining_consultants,
            "recurring_issues": recurring_issues,
            "batch_objects": batches,
        }

    def _group_by_week(self, batches: list) -> list:
        from datetime import timedelta
        weeks = defaultdict(list)
        for b in batches:
            week_start = b.start_time - timedelta(days=b.start_time.weekday())
            week_key = week_start.strftime("%Y-W%W")
            weeks[week_key].append(b)

        result = []
        for week_key, week_batches in sorted(weeks.items()):
            all_results = []
            for b in week_batches:
                all_results.extend(b.results)
            if not all_results:
                continue
            total = len(all_results)
            avg_score = sum(r.total_score for r in all_results) / total
            low_count = sum(1 for r in all_results if r.is_low_score)

            store_scores = defaultdict(list)
            for r in all_results:
                store_scores[r.recording.store].append(r.total_score)
            store_avgs = {s: round(sum(v) / len(v), 1) for s, v in store_scores.items()}

            result.append({
                "week": week_key,
                "batch_count": len(week_batches),
                "total_recordings": total,
                "avg_score": round(avg_score, 1),
                "low_count": low_count,
                "store_avgs": store_avgs,
            })
        return result

    def _find_risk_stores(self, batches: list) -> list:
        if len(batches) < 2:
            first = batches[0] if batches else None
            if not first:
                return []
            stores = self.analyzer.get_store_summaries(first)
            return [{"store": s.store_name, "avg_score": s.avg_score,
                      "reason": f"均分仅{s.avg_score}分"}
                     for s in stores if s.avg_score < self.low_score_threshold]

        all_store_data = defaultdict(list)
        for b in batches:
            for s in self.analyzer.get_store_summaries(b):
                all_store_data[s.store_name].append(s)

        risk = []
        for store, summaries in all_store_data.items():
            if len(summaries) < 2:
                continue
            sorted_s = sorted(summaries, key=lambda x: x.avg_score)
            worst = sorted_s[0]
            recent = summaries[-1]
            trend = recent.avg_score - sorted_s[0].avg_score

            reasons = []
            if worst.avg_score < self.low_score_threshold:
                reasons.append(f"均分低至{worst.avg_score}分")
            if trend < -5:
                reasons.append(f"趋势下降{abs(round(trend, 1))}分")
            if worst.issue_count > 5:
                reasons.append(f"问题数{worst.issue_count}个")

            if reasons:
                risk.append({
                    "store": store,
                    "avg_score": recent.avg_score,
                    "worst_score": worst.avg_score,
                    "trend": round(trend, 1),
                    "reasons": reasons,
                    "top_issues": worst.top_issues[:3],
                })

        risk.sort(key=lambda x: x["avg_score"])
        return risk[:10]

    def _find_declining_consultants(self, batches: list) -> list:
        if len(batches) < 2:
            return []

        all_cons_data = defaultdict(list)
        for b in batches:
            for c in self.analyzer.get_consultant_summaries(b):
                key = f"{c['store']}|{c['consultant']}"
                all_cons_data[key].append(c)

        declining = []
        for key, summaries in all_cons_data.items():
            if len(summaries) < 2:
                continue
            store, consultant = key.split("|", 1)
            sorted_s = sorted(summaries, key=lambda x: x["avg_score"])
            latest = summaries[-1]
            trend = latest["avg_score"] - sorted_s[0]["avg_score"]

            if trend < -5 or latest["low_count"] > 2:
                declining.append({
                    "store": store,
                    "consultant": consultant,
                    "latest_score": latest["avg_score"],
                    "trend": round(trend, 1),
                    "low_count": latest["low_count"],
                    "top_issues": latest["top_issues"][:3],
                })

        declining.sort(key=lambda x: x["trend"])
        return declining[:10]

    def _find_recurring_issues(self, batches: list) -> list:
        issue_by_batch = []
        for b in batches:
            top = dict(self.analyzer.get_top_issues(b, top_n=20))
            issue_by_batch.append(top)

        if not issue_by_batch:
            return []

        issue_counts = defaultdict(int)
        issue_total = defaultdict(int)
        for batch_issues in issue_by_batch:
            for issue, count in batch_issues.items():
                issue_counts[issue] += 1
                issue_total[issue] += count

        batch_count = len(batches)
        recurring = []
        for issue, appear_count in issue_counts.items():
            if appear_count >= max(2, batch_count * 0.5):
                recurring.append({
                    "issue": issue,
                    "appear_in_batches": appear_count,
                    "total_count": issue_total[issue],
                    "avg_per_batch": round(issue_total[issue] / batch_count, 1),
                    "frequency": f"{appear_count}/{batch_count}批次",
                })

        recurring.sort(key=lambda x: x["total_count"], reverse=True)
        return recurring[:10]

    def export_weekly_package(self, weekly_result: dict, exporter: ExcelExporter,
                               batch_a: BatchRecord, batch_b: BatchRecord) -> str:
        filepath = exporter.export_comparison(batch_a, batch_b, self.analyzer)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        ws_week = wb.create_sheet("周趋势")
        self._fill_week_trend_sheet(ws_week, weekly_result)

        ws_risk = wb.create_sheet("风险门店")
        self._fill_risk_stores_sheet(ws_risk, weekly_result)

        ws_decline = wb.create_sheet("退步咨询师")
        self._fill_declining_consultants_sheet(ws_decline, weekly_result)

        ws_recur = wb.create_sheet("反复问题")
        self._fill_recurring_issues_sheet(ws_recur, weekly_result)

        wb.save(filepath)
        return filepath

    def _fill_week_trend_sheet(self, ws, weekly_result):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws.merge_cells("A1:E1")
        ws["A1"] = "周趋势汇总"
        ws["A1"].font = Font(bold=True, size=14)

        headers = ["周", "批次数", "录音数", "平均分", "低分数"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, week in enumerate(weekly_result.get("weeks", []), 3):
            ws.cell(row=i, column=1, value=week["week"])
            ws.cell(row=i, column=2, value=week["batch_count"])
            ws.cell(row=i, column=3, value=week["total_recordings"])
            ws.cell(row=i, column=4, value=week["avg_score"])
            ws.cell(row=i, column=5, value=week["low_count"])

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10

    def _fill_risk_stores_sheet(self, ws, weekly_result):
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        ws.merge_cells("A1:E1")
        ws["A1"] = "重点风险门店"
        ws["A1"].font = Font(bold=True, size=14, color="C00000")

        headers = ["门店", "最新均分", "最低均分", "趋势", "原因"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, rs in enumerate(weekly_result.get("risk_stores", []), 3):
            ws.cell(row=i, column=1, value=rs["store"])
            ws.cell(row=i, column=2, value=rs["avg_score"])
            ws.cell(row=i, column=3, value=rs["worst_score"])
            ws.cell(row=i, column=4, value=rs["trend"])
            ws.cell(row=i, column=5, value="; ".join(rs["reasons"]))

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 40

    def _fill_declining_consultants_sheet(self, ws, weekly_result):
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        ws.merge_cells("A1:F1")
        ws["A1"] = "退步咨询师"
        ws["A1"].font = Font(bold=True, size=14, color="C00000")

        headers = ["门店", "咨询师", "最新均分", "趋势", "低分数", "主要问题"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, dc in enumerate(weekly_result.get("declining_consultants", []), 3):
            ws.cell(row=i, column=1, value=dc["store"])
            ws.cell(row=i, column=2, value=dc["consultant"])
            ws.cell(row=i, column=3, value=dc["latest_score"])
            ws.cell(row=i, column=4, value=dc["trend"])
            ws.cell(row=i, column=5, value=dc["low_count"])
            issues_str = "; ".join([f"{iss}({cnt})" for iss, cnt in dc["top_issues"]])
            ws.cell(row=i, column=6, value=issues_str)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 40

    def _fill_recurring_issues_sheet(self, ws, weekly_result):
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

        ws.merge_cells("A1:E1")
        ws["A1"] = "反复出现的问题"
        ws["A1"].font = Font(bold=True, size=14, color="BF8F00")

        headers = ["排名", "问题类型", "出现批次", "总次数", "批次均次"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, ri in enumerate(weekly_result.get("recurring_issues", []), 3):
            ws.cell(row=i, column=1, value=i - 2)
            ws.cell(row=i, column=2, value=ri["issue"])
            ws.cell(row=i, column=3, value=ri["frequency"])
            ws.cell(row=i, column=4, value=ri["total_count"])
            ws.cell(row=i, column=5, value=ri["avg_per_batch"])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
